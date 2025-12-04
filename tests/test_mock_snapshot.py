"""
End-to-End Test (Mock Snapshot)
Validiert: Schema-Validierung, Scoring, Export
Keine API-Aufrufe, rein synthetische Daten.
"""

import pandas as pd
import pytest
from pathlib import Path

from src.utils.validation import ensure_schema
from src.scores import compute_scores
from src.export import export_snapshot


@pytest.fixture
def mock_snapshot_df() -> pd.DataFrame:
    """Erzeugt synthetische Testdaten für den Snapshot."""
    data = [
        {
            "id": "alpha",
            "symbol": "alp",
            "market_cap": 200_000_000,
            "total_volume": 5_000_000,
            "current_price": 0.8,
            "ath": 1.0,
            "price_change_percentage_30d_in_currency": 25.0,
        },
        {
            "id": "beta",
            "symbol": "bet",
            "market_cap": 350_000_000,
            "total_volume": 12_000_000,
            "current_price": 0.5,
            "ath": 0.9,
            "price_change_percentage_30d_in_currency": -5.0,
        },
    ]
    return pd.DataFrame(data)


def test_mock_snapshot_end_to_end(tmp_path: Path, mock_snapshot_df: pd.DataFrame):
    """End-to-end Test: Schema -> Scoring -> Export"""

    schema_map = {
        "id": str,
        "symbol": str,
        "market_cap": float,
        "total_volume": float,
        "early_score": float,
        "breakout_score": float,
        "mexc_pair": str,
        "ath": float,
        "current_price": float,
    }

    # --- Step 1: Ensure Schema ---
    df = ensure_schema(mock_snapshot_df, schema_map)
    assert all(col in df.columns for col in schema_map.keys())

    # --- Step 2: Compute Scores ---
    df_scored = compute_scores(df)
    assert "early_score" in df_scored.columns
    assert df_scored["early_score"].notnull().all(), "Scores must be filled"

    # --- Step 3: Export to Excel ---
    export_path = tmp_path / "mock_snapshot.xlsx"
    export_snapshot(df_scored, export_path)

    # --- Step 4: Validate Output ---
    assert export_path.exists(), "Excel export file should exist"

    # read back basic metadata
    import openpyxl
    wb = openpyxl.load_workbook(export_path, read_only=True)
    assert "FullData" in wb.sheetnames or len(wb.sheetnames) > 0

    print(f"✅ Mock snapshot export successful: {export_path}")


def test_ensure_schema_handles_missing_columns():
    """Prüft, ob fehlende Spalten automatisch ergänzt werden."""
    df = pd.DataFrame({"id": ["x"], "symbol": ["y"]})
    schema_map = {"id": str, "symbol": str, "market_cap": float, "total_volume": float}

    df2 = ensure_schema(df, schema_map)

    assert "market_cap" in df2.columns
    assert "total_volume" in df2.columns
    assert df2["market_cap"].dtype in ("float64", "float32")

    print("✅ Missing column handling verified")
